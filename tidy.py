import pandas as pd
from openpyxl import Workbook
import numpy as np
import tkinter as tk
from tkinter import filedialog, messagebox
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches

def select_file():
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename()
    print(f"Selected file: {file_path}")
    return file_path


def get_processing_choice():
    def on_button_click():
        root.quit()

    root = tk.Tk()
    root.title("Processing Choice")
    var = tk.StringVar(value="entire")  # Default value is "entire"

    tk.Label(root, text="Choose processing option:").pack(padx=20, pady=10)
    tk.Radiobutton(root, text="Process entire file", variable=var, value="entire").pack()
    tk.Radiobutton(root, text="Extract specific soil type", variable=var, value="specific").pack()

    tk.Button(root, text="Confirm", command=on_button_click).pack(pady=10)

    root.mainloop()
    choice = var.get()  # Get the selected value after the loop ends
    root.destroy()
    return choice

def get_soil_type():
    def on_confirm():
        nonlocal soil_type  # Ensure we capture the value properly
        soil_type = entry.get()  # Get the value from the entry widget
        print('Soil type input:', soil_type)  # Debugging print
        root.quit()
    
    root = tk.Tk()
    root.title("Select Soil Type")
    
    # Initialize soil_type as a plain string
    soil_type = ""
    tk.Label(root, text="Enter the soil type you want to extract:").pack(padx=20, pady=10)
    entry = tk.Entry(root)
    entry.pack(padx=20, pady=10)
    
    tk.Button(root, text="Confirm", command=on_confirm).pack(pady=10)
    
    # Ensure the window appears in the foreground
    root.attributes('-topmost', True)
    root.update()
    root.attributes('-topmost', False)
    
    root.mainloop()  # Wait for the user to click "Confirm"
    
    # Use the string value from the entry
    result = soil_type
    print('Result after closing window:', result)  # Debugging print
    root.destroy()
    return result







# Get processing choice
processing_choice = get_processing_choice()
print(f"Processing choice: {processing_choice}")

# if processing_choice=="specific":
#     # Test the function in isolation
#     soil_type = get_soil_type()
#     print(f"Selected soil type: {soil_type}")

# Select file
selected_file = select_file()

# Read the Excel file
df = pd.read_excel(selected_file, header=0)

# Create a new workbook and select the active worksheet
new_wb = Workbook()
new_ws = new_wb.active

# Assuming your data is in the 1st, 2nd, and 13th columns (index 0, 1, and 12)
depth = df.iloc[:, 0].dropna()  # Clean depth values
depth = np.round(depth, 2)  # Round to 2 decimal places
value = df.iloc[:, 1].dropna()  # Clean value column
qc_dict = dict(zip(depth, value))
type_values = df.iloc[:, 18].dropna()
type_dict = dict(zip(depth, type_values))

# Get the number of rows in the DataFrame
num_points = int(max(depth)/0.02)

# Write numbers in column A and lookup values in column B and C
for i in range(1, num_points + 1):
    depth_value = round(0.02*i, 2)  # Calculate depth value
    new_ws.cell(row=i, column=1, value=depth_value)  # Write depth in the 1st column
    if depth_value in qc_dict:
        new_ws.cell(row=i, column=2, value=qc_dict[depth_value])  # Write value in the 2nd column
    if depth_value in type_dict:
        new_ws.cell(row=i, column=3, value=type_dict[depth_value])  # Write type in the 3rd column

# 将工作表中的数据读取到pandas DataFrame中
new_df = pd.DataFrame(new_ws.values)
#------------------------------------------------
# # 对缺失值进行插值
# new_df[1] = new_df[1].interpolate()
# new_df[2] = new_df[2].ffill()

# # 将插值后的数据写回到工作表
# for i in range(1, num_points + 1):
#     new_ws.cell(row=i, column=2, value=new_df.iloc[i-1, 1])
#     new_ws.cell(row=i, column=3, value=new_df.iloc[i-1, 2])
#------------------------------------------------
# 如果选择了特定土壤类型，则只保留该类型的行
if processing_choice == "specific":
    soil_type = get_soil_type()
    print('Selected soil type:', soil_type)
    
    for row in range(1, num_points + 1):
        cell_value = new_ws.cell(row=row, column=3).value  # Get the value in column C
        if cell_value is None:
            continue  # Skip rows where the cell is empty

        if str(cell_value) != soil_type:  # Compare as strings
            print(f"Row {row}: Deleting cell with value {cell_value} because it doesn't match {soil_type}")
            # Set the value of the cell in column C to None to delete the content
            new_ws.cell(row=row, column=1).value = None
            new_ws.cell(row=row, column=2).value = None


# Save the modified file
new_wb.save('modified_file.xlsx')
print('File saved as modified_file.xlsx')

# 畫圖
plot_depth=new_df[0]
plot_value=new_df[1]
plot_type=new_df[2]
print(plot_type)

# 定義顏色函數，根據條件返回顏色
def get_color(plot_type):
    if plot_type==1:
        return 'red'
    elif plot_type==2:
        return 'orange'
    elif plot_type ==3:
        return'green'
    elif plot_type ==4:
        return'blue'
    else:
        return 'black'
    
# 创建一个空白图
fig, ax = plt.subplots(figsize=(6, 12))
# x軸最大值為10
plt.xlim(0, 70)
# y軸最大值為110
plt.ylim(0, 110)


# 根据条件绘制线条
for i in range(len(plot_type)-1):
    color = get_color(plot_type[i])
    print(color)
    ax.plot(plot_value[i:i+2], plot_depth[i:i+2], color=color, lw=2)
# 反轉y軸
plt.gca().invert_yaxis() 


# 添加图例
red_patch = mpatches.Patch(color='red', label='Type 1')
orange_patch = mpatches.Patch(color='orange', label='Type 2')
green_patch = mpatches.Patch(color='green', label='Type 3')
blue_patch = mpatches.Patch(color='blue', label='Type 4')
black_patch = mpatches.Patch(color='black', label='Type 5')
plt.legend(handles=[red_patch, orange_patch, green_patch, blue_patch, black_patch])
plt.xlabel('qc')
plt.ylabel('Depth(m)')
plt.title('04 qc and soil type')
plt.grid(linestyle='--', linewidth=0.5)
x_major_locator = plt.MultipleLocator(5)
y_major_locator = plt.MultipleLocator(2)
ax.xaxis.set_major_locator(x_major_locator)
ax.yaxis.set_major_locator(y_major_locator)

# 先保存圖片
plt.savefig('04 qc and soil type.png')

# 然後顯示圖片
plt.show()
